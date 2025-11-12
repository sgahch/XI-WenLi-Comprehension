#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
学生批量导入测试Excel文件生成脚本

使用方法：
    python generate_test_excel.py

生成文件：
    - 学生批量导入测试数据_正常.xlsx
    - 学生批量导入测试数据_验证失败.xlsx
    - 学生批量导入测试数据_部门不存在.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def create_excel_template():
    """创建Excel模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生数据"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    
    # 表头
    headers = [
        '专业名称',
        '班级',
        '姓名',
        '学号（或教职工号）',
        '手机号码',
        '用户性别',
        '身份（学生、班委、辅导员）',
        '邮箱',
        '政治面貌',
        '入学日期'
    ]
    
    # 设置表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    return wb, ws

def generate_normal_test_data():
    """生成正常测试数据"""
    wb, ws = create_excel_template()
    
    # 测试数据
    test_data = [
        ['软件工程', '2401', '张三', '2507240101', '13800138001', '男', '学生', 'zhangsan@example.com', '团员', '2024-09-01'],
        ['软件工程', '2401', '李四', '2507240102', '13800138002', '女', '班委', 'lisi@example.com', '党员', '2024-09-01'],
        ['软件工程', '2402', '王五', '2507240201', '13800138003', '男', '学生', 'wangwu@example.com', '群众', '2024-09-01'],
        ['计算机科学与技术', '2401', '赵六', '2507230101', '13800138004', '女', '学生', 'zhaoliu@example.com', '团员', '2024-09-01'],
        ['计算机科学与技术', '2401', '孙七', '2507230102', '13800138005', '男', '班委', 'sunqi@example.com', '党员', '2024-09-01'],
        ['软件工程', '2401', '周八', '2507240103', '13800138006', '女', '学生', 'zhouba@example.com', '团员', '2024-09-01'],
        ['软件工程', '2402', '吴九', '2507240202', '13800138007', '男', '学生', 'wujiu@example.com', '群众', '2024-09-01'],
        ['计算机科学与技术', '2401', '郑十', '2507230103', '13800138008', '女', '学生', 'zhengshi@example.com', '团员', '2024-09-01'],
        ['软件工程', '2401', '周老师', 'T20240001', '13900139001', '女', '辅导员', 'zhoulaoshi@example.com', '党员', '2024-09-01'],
        ['计算机科学与技术', '2401', '陈老师', 'T20240002', '13900139002', '男', '辅导员', 'chenlaoshi@example.com', '党员', '2024-09-01'],
    ]
    
    # 填充数据
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 保存文件
    filename = '学生批量导入测试数据_正常.xlsx'
    wb.save(filename)
    print(f'✅ 已生成: {filename}')
    return filename

def generate_validation_fail_data():
    """生成验证失败测试数据"""
    wb, ws = create_excel_template()
    
    # 测试数据（包含各种验证错误）
    test_data = [
        # 学号为空
        ['软件工程', '2401', '测试1', '', '13800138010', '男', '学生', 'test1@example.com', '团员', '2024-09-01'],
        # 班级为空
        ['软件工程', '', '测试2', '2507240999', '13800138011', '女', '学生', 'test2@example.com', '团员', '2024-09-01'],
        # 专业名称为空
        ['', '2401', '测试3', '2507240998', '13800138012', '男', '学生', 'test3@example.com', '团员', '2024-09-01'],
        # 身份不合法
        ['软件工程', '2401', '测试4', '2507240997', '13800138013', '女', '教师', 'test4@example.com', '党员', '2024-09-01'],
        # 姓名为空
        ['软件工程', '2401', '', '2507240996', '13800138014', '男', '学生', 'test5@example.com', '团员', '2024-09-01'],
        # 身份为空
        ['软件工程', '2401', '测试6', '2507240995', '13800138015', '女', '', 'test6@example.com', '团员', '2024-09-01'],
    ]
    
    # 填充数据
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            # 标记错误单元格为红色
            if value == '' and col_num in [1, 2, 3, 4, 7]:  # 必填字段
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif col_num == 7 and value == '教师':  # 非法身份
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # 保存文件
    filename = '学生批量导入测试数据_验证失败.xlsx'
    wb.save(filename)
    print(f'✅ 已生成: {filename}')
    return filename

def generate_dept_not_found_data():
    """生成部门不存在测试数据"""
    wb, ws = create_excel_template()
    
    # 测试数据（部门不存在）
    test_data = [
        ['不存在的专业', '2401', '测试5', '2507249999', '13800138014', '男', '学生', 'test5@example.com', '团员', '2024-09-01'],
        ['软件工程', '9999', '测试6', '2507249998', '13800138015', '女', '学生', 'test6@example.com', '团员', '2024-09-01'],
    ]
    
    # 填充数据
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            # 标记不存在的部门为黄色
            if (col_num == 1 and value == '不存在的专业') or (col_num == 2 and value == '9999'):
                cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    
    # 保存文件
    filename = '学生批量导入测试数据_部门不存在.xlsx'
    wb.save(filename)
    print(f'✅ 已生成: {filename}')
    return filename

def generate_mixed_data():
    """生成混合测试数据（包含成功和失败的记录）"""
    wb, ws = create_excel_template()
    
    # 测试数据
    test_data = [
        # 正常数据
        ['软件工程', '2401', '张三', '2507240101', '13800138001', '男', '学生', 'zhangsan@example.com', '团员', '2024-09-01'],
        ['软件工程', '2401', '李四', '2507240102', '13800138002', '女', '班委', 'lisi@example.com', '党员', '2024-09-01'],
        # 学号为空（失败）
        ['软件工程', '2401', '测试1', '', '13800138010', '男', '学生', 'test1@example.com', '团员', '2024-09-01'],
        # 正常数据
        ['计算机科学与技术', '2401', '王五', '2507230101', '13800138003', '男', '学生', 'wangwu@example.com', '群众', '2024-09-01'],
        # 部门不存在（失败）
        ['不存在的专业', '2401', '测试2', '2507249999', '13800138014', '男', '学生', 'test2@example.com', '团员', '2024-09-01'],
        # 正常数据
        ['软件工程', '2402', '赵六', '2507240201', '13800138004', '女', '学生', 'zhaoliu@example.com', '团员', '2024-09-01'],
    ]
    
    # 填充数据
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 保存文件
    filename = '学生批量导入测试数据_混合.xlsx'
    wb.save(filename)
    print(f'✅ 已生成: {filename}')
    return filename

def main():
    """主函数"""
    print('=' * 60)
    print('学生批量导入测试Excel文件生成工具')
    print('=' * 60)
    print()
    
    try:
        # 生成各种测试文件
        generate_normal_test_data()
        generate_validation_fail_data()
        generate_dept_not_found_data()
        generate_mixed_data()
        
        print()
        print('=' * 60)
        print('✅ 所有测试文件生成完成！')
        print('=' * 60)
        print()
        print('📝 测试说明：')
        print('1. 学生批量导入测试数据_正常.xlsx - 用于测试正常导入流程')
        print('2. 学生批量导入测试数据_验证失败.xlsx - 用于测试数据验证')
        print('3. 学生批量导入测试数据_部门不存在.xlsx - 用于测试部门匹配')
        print('4. 学生批量导入测试数据_混合.xlsx - 用于测试混合场景')
        print()
        print('⚠️  注意：')
        print('- 请确保数据库中存在对应的部门（软件工程、计算机科学与技术）')
        print('- 请确保部门下存在对应的班级（2401、2402）')
        print('- 默认密码为：123456（可在sys_config表中配置）')
        print()
        
    except Exception as e:
        print(f'❌ 生成失败: {str(e)}')
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()

